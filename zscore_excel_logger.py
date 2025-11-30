"""
Z-Score Imbalance Iceberg Hunter Excel Logger

Comprehensive Excel logging for:
1. All trade entries/exits with P&L calculations
2. Strategy parameter calculations (imbalance, wall strength, delta Z-score, touches)
3. TP/SL and margin calculations
"""

import logging
from datetime import datetime
from typing import Dict, Optional, List
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)


class ZScoreExcelLogger:
    """
    Logs all Z-Score trades, parameters, and P&L to Excel.

    Sheets:
    - Trades
    - Parameters
    - Daily Summary
    """

    def __init__(self, filepath: Optional[str] = None) -> None:
        self.filepath = Path(filepath or config.EXCEL_LOG_FILE)
        self._init_workbook()

    # ======================================================================
    # Workbook and sheet initialization
    # ======================================================================

    def _init_workbook(self) -> None:
        if self.filepath.exists():
            logger.info(f"Opening existing Excel log: {self.filepath}")
            self.wb = load_workbook(str(self.filepath))
        else:
            logger.info(f"Creating new Excel log: {self.filepath}")
            self.wb = Workbook()
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

        if "Trades" not in self.wb.sheetnames:
            self._init_trades_sheet()
        if "Parameters" not in self.wb.sheetnames:
            self._init_parameters_sheet()
        if "Daily Summary" not in self.wb.sheetnames:
            self._init_summary_sheet()

        self.wb.save(str(self.filepath))
        logger.info(f"Excel logger initialized: {self.filepath}")

    def _init_trades_sheet(self) -> None:
        ws = self.wb.create_sheet("Trades")
        headers = [
            "Trade ID",
            "Entry Time",
            "Exit Time",
            "Duration (min)",
            "Side",
            "Entry Price",
            "Exit Price",
            "Quantity (BTC)",
            "Margin Used (USDT)",
            "Leverage",
            "Notional (USDT)",
            "TP Price",
            "SL Price",
            "Entry Imbalance",
            "Entry Z-Score",
            "Entry Wall Volume",
            "Entry HTF Trend",  # NEW
            "Exit Reason",
            "P&L (USDT)",
            "P&L %",
            "ROI % (on margin)",
            "Cumulative P&L (USDT)",
        ]
        ws.append(headers)
        self._style_header_row(ws, len(headers))

        widths = [
            16, 19, 19, 14, 8, 14, 14, 16, 16, 10, 16,
            14, 14, 16, 16, 18, 16, 20, 14, 10, 16, 20,
        ]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _init_parameters_sheet(self) -> None:
        ws = self.wb.create_sheet("Parameters")
        headers = [
            "Timestamp",
            "Current Price",
            "Imbalance",
            "Long Imb OK",
            "Short Imb OK",
            "Bid Wall Str",
            "Ask Wall Str",
            "Long Wall OK",
            "Short Wall OK",
            "Delta",
            "Z-Score",
            "Long Z OK",
            "Short Z OK",
            "Nearest Bid",
            "Nearest Ask",
            "Bid Dist Ticks",
            "Ask Dist Ticks",
            "Long Touch OK",
            "Short Touch OK",
            "HTF Trend",  # NEW
            "Decision",
            "Reason",
        ]
        ws.append(headers)
        self._style_header_row(ws, len(headers))

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    def _init_summary_sheet(self) -> None:
        ws = self.wb.create_sheet("Daily Summary")
        headers = [
            "Date",
            "Total Trades",
            "Winning Trades",
            "Losing Trades",
            "Win Rate %",
            "Total P&L (USDT)",
            "Max Win (USDT)",
            "Max Loss (USDT)",
            "Avg Win (USDT)",
            "Avg Loss (USDT)",
            "Profit Factor",
            "Avg Hold Time (min)",
        ]
        ws.append(headers)
        self._style_header_row(ws, len(headers))

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    # ======================================================================
    # Styling helpers
    # ======================================================================

    @staticmethod
    def _style_header_row(ws, num_cols: int) -> None:
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    @staticmethod
    def _style_data_row(ws, row: int, num_cols: int, is_winning: Optional[bool] = None) -> None:
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if is_winning is True:
            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            font = Font(color="006100")
        elif is_winning is False:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            font = Font(color="9C0006")
        else:
            fill = None
            font = Font()

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if fill:
                cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ======================================================================
    # Trade logging
    # ======================================================================

    def log_trade(
        self,
        trade_id: str,
        entry_time: str,
        exit_time: str,
        duration_minutes: float,
        side: str,
        entry_price: float,
        exit_price: float,
        quantity: float,
        margin_used: float,
        leverage: int,
        tp_price: float,
        sl_price: float,
        entry_imbalance: float,
        entry_z_score: float,
        entry_wall_volume: float,
        exit_reason: str,
        pnl_usdt: float,
        entry_htf_trend: str = "UNKNOWN",  # NEW
    ) -> None:
        try:
            ws = self.wb["Trades"]

            notional = entry_price * quantity
            pnl_pct = (pnl_usdt / notional * 100.0) if notional > 0 else 0.0
            roi_pct = (pnl_usdt / margin_used * 100.0) if margin_used > 0 else 0.0
            cumulative = self._get_cumulative_pnl() + pnl_usdt

            row = [
                trade_id,
                entry_time,
                exit_time,
                f"{duration_minutes:.1f}",
                side.upper(),
                f"{entry_price:.2f}",
                f"{exit_price:.2f}",
                f"{quantity:.6f}",
                f"{margin_used:.4f}",
                leverage,
                f"{notional:.2f}",
                f"{tp_price:.2f}",
                f"{sl_price:.2f}",
                f"{entry_imbalance:.3f}",
                f"{entry_z_score:.2f}",
                f"{entry_wall_volume:.4f}",
                entry_htf_trend,  # NEW
                exit_reason,
                f"{pnl_usdt:.2f}",
                f"{pnl_pct:.2f}",
                f"{roi_pct:.2f}",
                f"{cumulative:.2f}",
            ]

            ws.append(row)
            self._style_data_row(ws, ws.max_row, len(row), is_winning=(pnl_usdt > 0))
            self.wb.save(str(self.filepath))

            logger.info(
                f"[Excel] Trade {trade_id} | side={side} | htf={entry_htf_trend} | "
                f"pnl={pnl_usdt:.2f} USDT | roi={roi_pct:.2f}%"
            )

        except Exception as e:
            logger.error(f"Error logging trade: {e}")

    # ======================================================================
    # Parameter logging
    # ======================================================================

    def log_parameters(
        self,
        timestamp: str,
        current_price: float,
        imbalance_data: Dict,
        wall_data: Dict,
        delta_data: Dict,
        touch_data: Dict,
        decision: str,
        reason: str,
        htf_trend: Optional[str] = None,  # NEW
    ) -> None:
        try:
            ws = self.wb["Parameters"]

            row = [
                timestamp,
                f"{current_price:.2f}",
                f"{imbalance_data.get('imbalance', 0):.3f}",
                "YES" if imbalance_data.get("long_ok", False) else "NO",
                "YES" if imbalance_data.get("short_ok", False) else "NO",
                f"{wall_data.get('bid_wall_strength', 0):.2f}",
                f"{wall_data.get('ask_wall_strength', 0):.2f}",
                "YES" if wall_data.get("long_wall_ok", False) else "NO",
                "YES" if wall_data.get("short_wall_ok", False) else "NO",
                f"{delta_data.get('delta', 0):.2f}",
                f"{delta_data.get('z_score', 0):.2f}",
                "YES" if delta_data.get("long_ok", False) else "NO",
                "YES" if delta_data.get("short_ok", False) else "NO",
                f"{touch_data.get('nearest_bid', 0):.2f}",
                f"{touch_data.get('nearest_ask', 0):.2f}",
                f"{touch_data.get('bid_distance_ticks', 0):.2f}",
                f"{touch_data.get('ask_distance_ticks', 0):.2f}",
                "YES" if touch_data.get("long_touch_ok", False) else "NO",
                "YES" if touch_data.get("short_touch_ok", False) else "NO",
                htf_trend or "UNKNOWN",  # NEW
                decision,
                reason,
            ]

            ws.append(row)
            self._style_data_row(ws, ws.max_row, len(row))

            # Save periodically to reduce I/O pressure
            if ws.max_row % 50 == 0:
                self.wb.save(str(self.filepath))

        except Exception as e:
            logger.error(f"Error logging parameters: {e}")

    # ======================================================================
    # Aggregate stats
    # ======================================================================

    def _get_cumulative_pnl(self) -> float:
        try:
            ws = self.wb["Trades"]
            total = 0.0
            # P&L column is now index 19 (was 18 before adding HTF Trend)
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=19).value
                try:
                    total += float(val)
                except (TypeError, ValueError):
                    pass
            return total
        except Exception:
            return 0.0

    def update_daily_summary(self) -> None:
        try:
            ws_trades = self.wb["Trades"]
            ws_sum = self.wb["Daily Summary"]

            if ws_trades.max_row < 2:
                return

            pnls: List[float] = []
            durations: List[float] = []

            for row in range(2, ws_trades.max_row + 1):
                try:
                    # P&L column is now 19
                    pnls.append(float(ws_trades.cell(row=row, column=19).value or 0))
                    durations.append(float(ws_trades.cell(row=row, column=4).value or 0))
                except (TypeError, ValueError):
                    pass

            if not pnls:
                return

            total_trades = len(pnls)
            wins = [p for p in pnls if p > 0]
            losses = [p for p in pnls if p < 0]

            winning_trades = len(wins)
            losing_trades = len(losses)
            win_rate = winning_trades / total_trades * 100.0 if total_trades else 0.0

            total_pnl = sum(pnls)
            max_win = max(wins) if wins else 0.0
            max_loss = min(losses) if losses else 0.0

            avg_win = sum(wins) / len(wins) if wins else 0.0
            avg_loss = sum(losses) / len(losses) if losses else 0.0

            profit_factor = abs(avg_win / avg_loss) if avg_loss != 0 else 0.0
            avg_hold = sum(durations) / len(durations) if durations else 0.0

            date_str = datetime.now().strftime("%Y-%m-%d")

            # Update today's row if exists, otherwise append
            existing_row = None
            for r in range(2, ws_sum.max_row + 1):
                if (ws_sum.cell(row=r, column=1).value or "") == date_str:
                    existing_row = r
                    break

            row_values = [
                date_str,
                total_trades,
                winning_trades,
                losing_trades,
                f"{win_rate:.1f}%",
                f"{total_pnl:.2f}",
                f"{max_win:.2f}",
                f"{max_loss:.2f}",
                f"{avg_win:.2f}",
                f"{avg_loss:.2f}",
                f"{profit_factor:.2f}",
                f"{avg_hold:.1f}",
            ]

            if existing_row:
                for c, val in enumerate(row_values, 1):
                    ws_sum.cell(row=existing_row, column=c, value=val)
                self._style_data_row(ws_sum, existing_row, len(row_values))
            else:
                ws_sum.append(row_values)
                self._style_data_row(ws_sum, ws_sum.max_row, len(row_values))

            self.wb.save(str(self.filepath))

        except Exception as e:
            logger.error(f"Error updating daily summary: {e}")

    def close(self) -> None:
        try:
            self.update_daily_summary()
            self.wb.save(str(self.filepath))
            logger.info(f"Excel log saved: {self.filepath}")
        except Exception as e:
            logger.error(f"Error closing Excel log: {e}")
